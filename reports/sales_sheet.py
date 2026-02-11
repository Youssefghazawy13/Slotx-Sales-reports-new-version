from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def auto_fit(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 3


def create_sales_sheet(wb, brand_sales, mode):
    ws = wb.create_sheet("Sales")

    # ====== COLORS (نفس inventory) ======
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    row_fill_1 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF")
    normal_font = Font(bold=False)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Branch",
        "Brand",
        "Product",
        "Barcode",
        "Quantity",
        "Total Price"
    ]

    ws.append(headers)

    # ===== HEADER STYLE =====
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    total_qty = 0
    total_money = 0
    current_row = 2

    # ===== DATA ROWS =====
    for index, row in brand_sales.iterrows():
        qty = row.get("quantity", 0)
        price = row.get("total", 0)

        total_qty += qty
        total_money += price

        ws.append([
            mode,
            row.get("brand", ""),
            row.get("product_name", ""),
            row.get("barcode", ""),
            qty,
            f"{price:,.2f} EGP"
        ])

        fill = row_fill_1 if current_row % 2 == 0 else row_fill_2

        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = fill
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1

    # ===== TOTAL ROW (نفس ستايل الجدول) =====
    ws.append([
        "",
        "",
        "",
        "",
        f"Total={int(total_qty)}",
        f"Total={total_money:,.2f} EGP"
    ])

    fill = row_fill_1 if current_row % 2 == 0 else row_fill_2

    for col in range(1, 7):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    auto_fit(ws)