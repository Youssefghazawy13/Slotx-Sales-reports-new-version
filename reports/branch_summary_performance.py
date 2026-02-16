from openpyxl.styles import Font, PatternFill, Alignment
from utils.excel_helpers import auto_fit_columns
from core.deals_engine import normalize_brand_name


def create_performance_sheet(
    wb,
    branch_name,
    payout_cycle,
    sales_df,
    inventory_df,
    deals_dict
):

    ws = wb.create_sheet("Performance")

    # =====================================================
    # 🔥 IMPORTANT FIX — NORMALIZE BEFORE GROUPING
    # =====================================================

    sales_df = sales_df.copy()
    inventory_df = inventory_df.copy()

    sales_df["brand_original"] = sales_df["brand"]
    inventory_df["brand_original"] = inventory_df["brand"]

    sales_df["brand"] = sales_df["brand"].astype(str).apply(normalize_brand_name)
    inventory_df["brand"] = inventory_df["brand"].astype(str).apply(normalize_brand_name)

    # =====================================================
    # GROUP SALES PER BRAND
    # =====================================================

    sales_grouped = (
        sales_df.groupby("brand")
        .agg({
            "quantity": "sum",
            "total": "sum",
            "brand_original": "first"
        })
        .reset_index()
    )

    sales_grouped = sales_grouped.sort_values(
        by="total",
        ascending=False
    ).reset_index(drop=True)

    sales_grouped["Rank"] = sales_grouped.index + 1

    # =====================================================
    # GROUP INVENTORY PER BRAND
    # =====================================================

    inventory_grouped = (
        inventory_df.groupby("brand")
        .agg({
            "available_quantity": "sum",
            "sale_price": "mean"
        })
        .reset_index()
    )

    inventory_grouped["inventory_value"] = (
        inventory_grouped["available_quantity"] *
        inventory_grouped["sale_price"]
    )

    # =====================================================
    # MERGE ON NORMALIZED BRAND
    # =====================================================

    summary_df = sales_grouped.merge(
        inventory_grouped,
        on="brand",
        how="left"
    )

    summary_df["available_quantity"] = summary_df["available_quantity"].fillna(0)
    summary_df["inventory_value"] = summary_df["inventory_value"].fillna(0)

    # =====================================================
    # APPLY DEALS (SMART MATCHING)
    # =====================================================

    after_percentage_list = []
    after_rent_list = []
    after_all_list = []
    percentage_deduction_list = []
    rent_deduction_list = []

    for _, row in summary_df.iterrows():

        normalized_brand = row["brand"]
        sales_money = row["total"]

        deal = deals_dict.get(
            normalized_brand,
            {"percentage": 0, "rent": 0}
        )

        percentage = deal["percentage"]
        rent = deal["rent"]

        percentage_deduction = sales_money * (percentage / 100)
        after_percentage = sales_money - percentage_deduction
        after_rent = after_percentage - rent

        percentage_deduction_list.append(percentage_deduction)
        rent_deduction_list.append(rent)
        after_percentage_list.append(after_percentage)
        after_rent_list.append(after_rent)
        after_all_list.append(after_rent)

    summary_df["percentage_deduction"] = percentage_deduction_list
    summary_df["rent_deduction"] = rent_deduction_list
    summary_df["after_percentage"] = after_percentage_list
    summary_df["after_rent"] = after_rent_list
    summary_df["after_all"] = after_all_list

    # =====================================================
    # KPI CARDS
    # =====================================================

    total_sales_money = summary_df["total"].sum()
    total_percentage_deduction = summary_df["percentage_deduction"].sum()
    total_rent_deduction = summary_df["rent_deduction"].sum()
    total_after_all = summary_df["after_all"].sum()
    total_deductions = total_percentage_deduction + total_rent_deduction

    kpi_fill = PatternFill(
        start_color="0A1F5C",
        end_color="0A1F5C",
        fill_type="solid"
    )

    kpi_titles = [
        "Total Branch Sales",
        "Total % Deducted",
        "Total Rent Deducted",
        "Total Deductions",
        "Sales After All Deductions"
    ]

    kpi_values = [
        total_sales_money,
        total_percentage_deduction,
        total_rent_deduction,
        total_deductions,
        total_after_all
    ]

    for i, (title, value) in enumerate(zip(kpi_titles, kpi_values)):

        col_letter = chr(65 + i)

        ws[f"{col_letter}1"] = title
        ws[f"{col_letter}2"] = value

        ws[f"{col_letter}1"].fill = kpi_fill
        ws[f"{col_letter}2"].fill = kpi_fill

        ws[f"{col_letter}1"].font = Font(bold=True, color="FFFFFF")
        ws[f"{col_letter}2"].font = Font(bold=True, color="FFFFFF")

        ws[f"{col_letter}1"].alignment = Alignment(horizontal="center")
        ws[f"{col_letter}2"].alignment = Alignment(horizontal="center")

        ws[f"{col_letter}2"].number_format = '#,##0.00 "EGP"'
        ws.column_dimensions[col_letter].width = 26

    # =====================================================
    # PERFORMANCE TABLE
    # =====================================================

    ws.append([])
    ws.append([])

    headers = [
        "Rank",
        "Brand",
        "Sales Qty",
        "Sales Money",
        "After Percentage",
        "After Rent",
        "After All Deductions",
        "Inventory Qty",
        "Inventory Value"
    ]

    ws.append(headers)

    header_row = ws.max_row

    header_fill = PatternFill(
        start_color="0A1F5C",
        end_color="0A1F5C",
        fill_type="solid"
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for _, row in summary_df.iterrows():
        ws.append([
            row["Rank"],
            row["brand_original"],  # 🔥 show original name
            row["quantity"],
            row["total"],
            row["after_percentage"],
            row["after_rent"],
            row["after_all"],
            row["available_quantity"],
            row["inventory_value"]
        ])

    auto_fit_columns(ws)
