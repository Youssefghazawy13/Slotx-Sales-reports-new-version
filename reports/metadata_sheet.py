# reports/metadata_sheet.py

from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from utils.excel_helpers import auto_fit_columns


def create_metadata_sheet(
    wb,
    mode: str,
    payout_cycle: str
):

    ws = wb.create_sheet("Metadata")

    # -------------------------
    # SLOT-X BANNER
    # -------------------------

    # Merge area for banner
    ws.merge_cells("A1:H4")

    title_cell = ws["A1"]
    title_cell.value = "SLOT-X"

    # Dark blue background (same vibe as logo)
    banner_fill = PatternFill(
        start_color="0A1F5C",   # Deep royal blue
        end_color="0A1F5C",
        fill_type="solid"
    )

    for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=8):
        for cell in row:
            cell.fill = banner_fill

    # Metallic silver text color
    title_cell.font = Font(
        name="Impact",     # أقرب حاجة للسمك ده
        size=44,
        bold=True,
        color="D9D9D9"     # Silver metallic tone
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Increase height for strong visual
    for i in range(1, 5):
        ws.row_dimensions[i].height = 45

    # -------------------------
    # Metadata Section
    # -------------------------

    dark_blue_font = Font(color="1F4E78")
    bold_dark_blue_font = Font(bold=True, color="1F4E78")

    start_row = 6

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    metadata_rows = [
        ["Powered by:", "Slot-X Solutions"],
        ["Version:", "v1.0"],
        ["Report Type:", mode],
        ["Payout Cycle:", payout_cycle],
        ["Generated At:", now],
    ]

    current_row = start_row

    for row in metadata_rows:
        ws.cell(row=current_row, column=1, value=row[0]).font = bold_dark_blue_font
        ws.cell(row=current_row, column=2, value=row[1]).font = dark_blue_font
        current_row += 1

    auto_fit_columns(ws)
