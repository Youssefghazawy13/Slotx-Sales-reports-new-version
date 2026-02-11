from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

from reports.branch_summary_performance import create_performance_sheet
from reports.metadata_sheet import create_metadata_sheet
from reports.sales_sheet import create_sales_sheet
from reports.inventory_sheet import create_inventory_sheet

from utils.excel_helpers import auto_fit_columns


def build_branch_summary_workbook(
    branch_name,
    payout_cycle,
    sales_df,
    inventory_df,
    deals_dict
):

    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # =====================================================
    # ALL SALES
    # =====================================================

    create_sales_sheet(
        wb=wb,
        brand_sales=sales_df,
        mode=branch_name
    )

    # =====================================================
    # ALL INVENTORY
    # =====================================================

    create_inventory_sheet(
        wb=wb,
        brand_inventory=inventory_df,
        mode=branch_name
    )

    # =====================================================
    # DEALS TAB (Styled Like Other Tables)
    # =====================================================

    ws_deals = wb.create_sheet("Deals")

    headers = ["Brand", "Deal %", "Rent (EGP)"]
    ws_deals.append(headers)

    header_fill = PatternFill(
        start_color="0A1F5C",
        end_color="0A1F5C",
        fill_type="solid"
    )

    for col in range(1, 4):
        cell = ws_deals.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for brand, deal in deals_dict.items():
        ws_deals.append([
            brand,
            deal.get("percentage", 0),
            deal.get("rent", 0)
        ])

    last_row = ws_deals.max_row

    # Apply table style
    table = Table(
        displayName="DealsTable",
        ref=f"A1:C{last_row}"
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True
    )

    table.tableStyleInfo = style
    ws_deals.add_table(table)

    auto_fit_columns(ws_deals)

    # =====================================================
    # PERFORMANCE TAB (Renamed Properly)
    # =====================================================

    create_performance_sheet(
        wb=wb,
        branch_name=branch_name,
        payout_cycle=payout_cycle,
        sales_df=sales_df,
        inventory_df=inventory_df,
        deals_dict=deals_dict
    )

    # =====================================================
    # METADATA
    # =====================================================

    create_metadata_sheet(
        wb,
        f"{branch_name} Summary",
        branch_name,
        payout_cycle
    )

    return wb