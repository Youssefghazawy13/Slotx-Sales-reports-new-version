from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from utils.excel_helpers import auto_fit_columns


def detect_column(df, possible_names):

    for col in df.columns:
        if col.lower().strip() in possible_names:
            return col
    return None


def get_status(qty):

    if qty <= 2:
        return "Critical"
    elif qty <= 5:
        return "Low"
    elif qty <= 10:
        return "Medium"
    else:
        return "Good"


def create_inventory_sheet(wb, brand_inventory, mode):

    ws = wb.create_sheet("Inventory")

    if brand_inventory.empty:
        ws.append(["No Inventory Data"])
        return

    # =====================================================
    # Detect Columns Dynamically
    # =====================================================

    product_col = detect_column(
        brand_inventory,
        ["product_name", "product", "product name", "name", "name_en"]
    )

    price_col = detect_column(
        brand_inventory,
        ["price", "sale_price", "product price"]
    )

    barcode_col = detect_column(
        brand_inventory,
        ["barcode", "barcodes", "sku", "code"]
    )

    is_merged = mode.lower() == "merged"

    # =====================================================
    # Headers
    # =====================================================

    if is_merged:

        headers = [
            "Product",
            "Barcode",
            "Price",
            "Alexandria Qty",
            "Zamalek Qty",
            "Total Qty",
            "Status",
            "Notes"
        ]

    else:

        headers = [
            "Product",
            "Barcode",
            "Price",
            "Quantity",
            "Status",
            "Notes"
        ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    critical_count = 0

    # =====================================================
    # Write Rows
    # =====================================================

    for _, row in brand_inventory.iterrows():

        product = row.get(product_col, "") if product_col else ""
        barcode = row.get(barcode_col, "") if barcode_col else ""
        price = row.get(price_col, 0) if price_col else 0

        if is_merged:

            alex_qty = row.get("alex_qty", 0)
            zam_qty = row.get("zamalek_qty", 0)
            total_qty = alex_qty + zam_qty

            status = get_status(total_qty)

            ws.append([
                product,
                barcode,
                price,
                alex_qty,
                zam_qty,
                total_qty,
                status,
                ""
            ])

            qty_for_status = total_qty

        else:

            qty = row.get("quantity", 0)
            status = get_status(qty)

            ws.append([
                product,
                barcode,
                price,
                qty,
                status,
                ""
            ])

            qty_for_status = qty

        if qty_for_status <= 2:
            critical_count += 1

    # =====================================================
    # Notes Logic
    # =====================================================

    if critical_count >= 3:
        ws.cell(row=2, column=len(headers)).value = \
            "⚠ Brand requires urgent restocking"

    # =====================================================
    # Price Format
    # =====================================================

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '#,##0.00'

    # =====================================================
    # Color Status
    # =====================================================

    status_col_index = headers.index("Status") + 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):

        status_cell = row[status_col_index - 1]

        if status_cell.value == "Critical":
            status_cell.fill = PatternFill(start_color="FFC7CE",
                                           end_color="FFC7CE",
                                           fill_type="solid")

        elif status_cell.value == "Low":
            status_cell.fill = PatternFill(start_color="FFEB9C",
                                           end_color="FFEB9C",
                                           fill_type="solid")

        elif status_cell.value == "Medium":
            status_cell.fill = PatternFill(start_color="C6EFCE",
                                           end_color="C6EFCE",
                                           fill_type="solid")

        elif status_cell.value == "Good":
            status_cell.fill = PatternFill(start_color="A9D08E",
                                           end_color="A9D08E",
                                           fill_type="solid")

    # =====================================================
    # Excel Table Style
    # =====================================================

    table = Table(
        displayName="InventoryTable",
        ref=ws.dimensions
    )

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    auto_fit_columns(ws)
