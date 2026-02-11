from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from utils.excel_helpers import auto_fit_columns
from core.kpi_engine import apply_deal


def create_report_sheet(
    wb,
    brand_name: str,
    mode: str,
    payout_cycle: str,
    brand_sales,
    total_inventory_qty: float,
    total_inventory_value: float,
    total_sales_qty: float,
    total_sales_money: float,
    deals_dict: dict
):

    ws = wb.create_sheet("Report")

    percentage = deals_dict.get(brand_name, {}).get("percentage", 0)
    rent = deals_dict.get(brand_name, {}).get("rent", 0)

    after_percentage, after_rent = apply_deal(
        total_sales_money,
        percentage,
        rent
    )

    # =========================
    # FIX KPI COLUMN WIDTHS
    # =========================

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 5
    ws.column_dimensions["F"].width = 5
    ws.column_dimensions["G"].width = 22

    # =========================
    # KPI CARDS (ROW 1)
    # =========================

    def create_kpi_card(row, col_letter, title, value):

        cell = ws[f"{col_letter}{row}"]
        cell.value = f"{title}\n{value}"

        cell.font = Font(size=12, bold=True, color="FFFFFF")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        ws.row_dimensions[row].height = 45

        fill = PatternFill(
            start_color="0A1F5C",
            end_color="0A1F5C",
            fill_type="solid"
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        cell.fill = fill
        cell.border = border

    create_kpi_card(1, "B", "Total Sales",
                    f"{total_sales_money:,.2f} EGP")

    create_kpi_card(1, "D", "Net After Deal",
                    f"{after_rent:,.2f} EGP")

    create_kpi_card(1, "G", "Inventory Units",
                    total_inventory_qty)

    # =========================
    # REPORT DETAILS
    # =========================

    row_pointer = 4

    ws.cell(row=row_pointer, column=1,
            value="Branch Name:").font = Font(bold=True)
    ws.cell(row=row_pointer, column=2, value=mode)

    ws.cell(row=row_pointer+1, column=1,
            value="Brand Name:").font = Font(bold=True)
    ws.cell(row=row_pointer+1, column=2, value=brand_name)

    ws.cell(row=row_pointer+2, column=1,
            value="Payout Cycle:").font = Font(bold=True)
    ws.cell(row=row_pointer+2, column=2, value=payout_cycle)

    # Skip 2 rows
    row_pointer += 5

    ws.cell(row=row_pointer, column=1,
            value="Total Inventory Quantity:").font = Font(bold=True)
    ws.cell(row=row_pointer, column=2, value=total_inventory_qty)

    ws.cell(row=row_pointer+1, column=1,
            value="Total Inventory Value:").font = Font(bold=True)
    ws.cell(row=row_pointer+1, column=2,
            value=f"{total_inventory_value:,.2f} EGP")

    # Skip 2 rows
    row_pointer += 4

    # Best Selling Products (Top 2)
    product_sales = (
        brand_sales.groupby("product_name")["quantity"]
        .sum()
        .sort_values(ascending=False)
        .head(2)
    )

    best_products_text = ""
    for product, qty in product_sales.items():
        best_products_text += f"{product} ({qty})  "

    ws.cell(row=row_pointer, column=1,
            value="Best Selling Products:").font = Font(bold=True)
    ws.cell(row=row_pointer, column=2, value=best_products_text)

    # Best Selling Size
    best_size = ""
    if "size" in brand_sales.columns:
        size_sales = (
            brand_sales.groupby("size")["quantity"]
            .sum()
            .sort_values(ascending=False)
        )
        if not size_sales.empty:
            best_size = size_sales.index[0]

    ws.cell(row=row_pointer+1, column=1,
            value="Best Selling Size:").font = Font(bold=True)
    ws.cell(row=row_pointer+1, column=2, value=best_size)

    # Skip 2 rows
    row_pointer += 4

    ws.cell(row=row_pointer, column=1,
            value="Total Sales Quantity:").font = Font(bold=True)
    ws.cell(row=row_pointer, column=2, value=total_sales_qty)

    ws.cell(row=row_pointer+1, column=1,
            value="Total Sales Money:").font = Font(bold=True)
    ws.cell(row=row_pointer+1, column=2,
            value=f"{total_sales_money:,.2f} EGP")

    ws.cell(row=row_pointer+2, column=1,
            value="After Percentage:").font = Font(bold=True)
    ws.cell(row=row_pointer+2, column=2,
            value=f"{after_percentage:,.2f} EGP")

    ws.cell(row=row_pointer+3, column=1,
            value="After Rent:").font = Font(bold=True)
    ws.cell(row=row_pointer+3, column=2,
            value=f"{after_rent:,.2f} EGP")

    auto_fit_columns(ws)

    # Re-fix KPI columns after auto-fit
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["G"].width = 22
