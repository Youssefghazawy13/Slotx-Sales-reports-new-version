from openpyxl.styles import Font

def build_sales(wb, sales_df, brand, branch):
    ws = wb.create_sheet("Sales")

    headers = ["Branch", "Brand", "Product", "Barcode", "Quantity", "Price"]
    ws.append(headers)

    for c in ws[1]:
        c.font = Font(bold=True)

    total_qty = 0
    total_price = 0

    brand_sales = sales_df[sales_df["brand"] == brand]

    for _, r in brand_sales.iterrows():
        qty = r.get("quantity", 0)
        price = r.get("total", 0)

        ws.append([
            branch,
            brand,
            r.get("product"),   # ✅ اسم المنتج
            r.get("barcode"),
            qty,
            price
        ])

        total_qty += qty
        total_price += price

    # ✅ Total مرة واحدة فقط + Bold
    ws.append(["", "", "", "Total", total_qty, total_price])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
