# utils/excel_helpers.py

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# -------------------------
# AUTO FIT
# -------------------------

def auto_fit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass

        adjusted_width = max(12, min(max_length + 3, 50))
        ws.column_dimensions[column_letter].width = adjusted_width


# -------------------------
# HEADER STYLE
# -------------------------

def apply_header_style(ws):

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.freeze_panes = "A2"


# -------------------------
# STATUS COLORING
# -------------------------

def get_status_fill(status):

    colors = {
        "Low Stock": "FF4C4C",
        "Slow Moving": "FFA500",
        "No Deal": "9E9E9E",
        "Healthy": "2E7D32"
    }

    return PatternFill(
        start_color=colors.get(status, "FFFFFF"),
        end_color=colors.get(status, "FFFFFF"),
        fill_type="solid"
    )


# -------------------------
# MONEY FORMAT
# -------------------------

def format_money_cell(cell):
    cell.number_format = '#,##0.00'
